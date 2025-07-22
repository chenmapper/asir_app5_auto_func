import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# --- SUM 檢測 subtask (from v11) with numeric guard ---
def apply_span_formula_horizontal(df, ws, start_col, end_col, col_idx, yellow_fill):
    flags = []
    n_rows, _ = df.shape
    for r in range(n_rows):
        val = df.iat[r, col_idx]
        try:
            val_num = float(val)
        except:
            flags.append(False)
            continue
        segment = df.iloc[r, start_col:end_col+1]
        seg_vals = pd.to_numeric(segment, errors='coerce')
        seg_sum = seg_vals.sum()
        flags.append(abs(val_num - seg_sum) < 1e-6)
    count = 0; r = 0
    while r < n_rows:
        if flags[r]:
            start = r
            while r < n_rows and flags[r]: r += 1
            length = r - start
            if length >= 3:
                first = get_column_letter(start_col+1)
                last = get_column_letter(end_col+1)
                for i in range(length):
                    row = start + i + 2
                    ws.cell(row=row, column=col_idx+1).value = f"=SUM({first}{row}:{last}{row})"
                    ws.cell(row=row, column=col_idx+1).fill = yellow_fill
                    count += 1
        else:
            r += 1
    return count

def apply_span_formula_vertical(df, ws, start_row, end_row, row_idx, yellow_fill):
    _, n_cols = df.shape
    flags = []
    for c in range(n_cols):
        val = df.iat[row_idx, c]
        try:
            val_num = float(val)
        except:
            flags.append(False)
            continue
        segment = df.iloc[start_row:end_row+1, c]
        seg_vals = pd.to_numeric(segment, errors='coerce')
        seg_sum = seg_vals.sum()
        flags.append(abs(val_num - seg_sum) < 1e-6)
    count = 0; c = 0
    while c < n_cols:
        if flags[c]:
            start = c
            while c < n_cols and flags[c]: c += 1
            length = c - start
            if length >= 3:
                for i in range(length):
                    col_idx = start + i + 1
                    row = row_idx + 2
                    letter = get_column_letter(col_idx)
                    ws.cell(row=row, column=col_idx).value = f"=SUM({letter}{start_row+2}:{letter}{end_row+2})"
                    ws.cell(row=row, column=col_idx).fill = yellow_fill
                    count += 1
        else:
            c += 1
    return count

# --- VLOOKUP 生成 subtask (from v12) ---
def generate_vlookup(ws, df_key, key_sheet, key_col, subset_col, subset_val,
                     df_table, table_sheet, start_idx, output_cols, blue_fill):
    first = get_column_letter(start_idx+1)
    last = get_column_letter(len(df_table.columns))
    table_range = f"'{table_sheet}'!${first}$2:${last}${len(df_table)+1}"
    key_cols = df_key.columns.tolist()
    for i, col in enumerate(key_cols, start=1):
        ws.cell(row=1, column=i).value = col
    for j, col in enumerate(output_cols, start=len(key_cols)+1):
        ws.cell(row=1, column=j).value = col
    filtered = df_key if not subset_col else df_key[df_key[subset_col].astype(str)==subset_val]
    for r, tup in enumerate(filtered.itertuples(index=False), start=2):
        for i, val in enumerate(tup, start=1):
            ws.cell(row=r, column=i).value = val
        lookup = f"{get_column_letter(key_cols.index(key_col)+1)}{r}"
        for j, col in enumerate(output_cols, start=len(key_cols)+1):
            idx = df_table.columns.get_loc(col) - start_idx + 1
            cell = ws.cell(row=r, column=j)
            cell.value = f"=VLOOKUP({lookup},{table_range},{idx},FALSE)"
            cell.fill = blue_fill

# --- SUMIF 生成 subtask with dual-condition support ---
def generate_count_and_sum_sheet(res_ws, df, crit1, crit2=None, sum_col=None,
                                 sheet_name=None, start_col=1, filter_col=None, filter_val=None):
    green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    blue = PatternFill(start_color='CCECFF', end_color='CCECFF', fill_type='solid')
    df_sub = df[df[filter_col]==filter_val] if filter_col else df
    n = len(df)
    c1 = df.columns.get_loc(crit1)+1
    r1 = f"'{sheet_name}'!{get_column_letter(c1)}2:{get_column_letter(c1)}{n+1}"
    si = df.columns.get_loc(sum_col)+1
    sr = f"'{sheet_name}'!{get_column_letter(si)}2:{get_column_letter(si)}{n+1}"
    # Headers
    res_ws.cell(row=1, column=start_col).value = crit1
    if crit2:
        res_ws.cell(row=1, column=start_col+1).value = crit2
        res_ws.cell(row=1, column=start_col+2).value = 'Count'
        res_ws.cell(row=1, column=start_col+3).value = f'Sum({sum_col})'
        c2 = df.columns.get_loc(crit2)+1
        r2 = f"'{sheet_name}'!{get_column_letter(c2)}2:{get_column_letter(c2)}{n+1}"
        pairs = df_sub[[crit1, crit2]].drop_duplicates().dropna().values.tolist()
        for i, (v1, v2) in enumerate(pairs, start=2):
            res_ws.cell(row=i, column=start_col).value = v1
            res_ws.cell(row=i, column=start_col+1).value = v2
            ref1 = f"{get_column_letter(start_col)}{i}"
            ref2 = f"{get_column_letter(start_col+1)}{i}"
            cell_c = res_ws.cell(row=i, column=start_col+2)
            cell_c.value = f"=COUNTIFS({r1},{ref1},{r2},{ref2})"; cell_c.fill = green
            cell_s = res_ws.cell(row=i, column=start_col+3)
            cell_s.value = f"=SUMIFS({sr},{r1},{ref1},{r2},{ref2})"; cell_s.fill = blue
        return 4, len(pairs)
    else:
        res_ws.cell(row=1, column=start_col+1).value = 'Count'
        res_ws.cell(row=1, column=start_col+2).value = f'Sum({sum_col})'
        vals = sorted(df_sub[crit1].dropna().unique())
        for i, v in enumerate(vals, start=2):
            res_ws.cell(row=i, column=start_col).value = v
            ref = f"{get_column_letter(start_col)}{i}"
            cell_c = res_ws.cell(row=i, column=start_col+1)
            cell_c.value = f"=COUNTIF({r1},{ref})"; cell_c.fill = green
            cell_s = res_ws.cell(row=i, column=start_col+2)
            cell_s.value = f"=SUMIF({r1},{ref},{sr})"; cell_s.fill = blue
        return 3, len(vals)

# --- Multi-Column Link Subtask ---
def generate_link_sheet(wb, source_sheet, link_cols, df_src):
    title = f"link_{source_sheet}"
    if title in wb.sheetnames:
        return title
    ws_src = wb[source_sheet]
    ws = wb.create_sheet(title)
    for row in ws_src.iter_rows(values_only=True): ws.append(row)
    ws.insert_cols(1)
    combined_header = "_".join(link_cols)
    ws.cell(row=1, column=1).value = combined_header
    for r in range(2, ws.max_row+1):
        parts = [str(df_src.at[r-2, col]) if df_src.at[r-2, col] is not None else "" for col in link_cols]
        ws.cell(row=r, column=1).value = "_".join(parts)
    return title

# --- Main App ---
st.set_page_config(page_title='App5 通用轉公式', layout='wide')
st.title('App5 通用轉公式工具')
# --- File Upload & Flow Prompt ---
uploaded = st.file_uploader('上傳 .xlsx 檔案', type='xlsx')
if not uploaded:
    st.stop()
st.markdown("""
**操作流程：**
1. 上傳您的 Excel 檔 (.xlsx)。  
2. 左側選擇模式並填寫參數：  
   - **SUMIF 生成**：條件1/2、篩選、加總欄→生成。  
   - **VLOOKUP 生成**：查找表、比對表、起始欄、擷取欄→生成。  
   - **多欄合一**：來源表、2~4欄位→生成連結結果。  
   - **SUM 檢測**：一鍵檢測並插入 SUM 公式。  
3. 完成後按「下載結果」，獲取 `result_通用公式.xlsx`。  
""")

# --- Initialize Workbook ---
bytes_data = uploaded.read()
if 'wb' not in st.session_state or st.session_state.filename != uploaded.name:
    uploaded.seek(0)
    st.session_state.wb = load_workbook(BytesIO(bytes_data))
    st.session_state.filename = uploaded.name
wb = st.session_state.wb
sheets = wb.sheetnames
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
blue_fill = PatternFill(start_color='CCECFF', end_color='CCECFF', fill_type='solid')

# --- Mode Selection ---
mode = st.sidebar.selectbox('選擇模式', ['SUMIF 生成', 'VLOOKUP 生成', '多欄合一', 'SUM 檢測'])

# --- Mode Handlers ---
if mode == 'SUMIF 生成':
    source = [s for s in sheets if s != 'sumif_result']
    data_sheet = st.sidebar.selectbox('來源工作表', source)
    df = pd.read_excel(BytesIO(bytes_data), sheet_name=data_sheet)
    crit1 = st.sidebar.selectbox('條件1', df.columns.tolist())
    sel2 = st.sidebar.selectbox('條件2(可空)', [''] + df.columns.tolist())
    crit2 = sel2 or None
    fcol = st.sidebar.selectbox('篩選欄(可空)', [''] + df.columns.tolist())
    fval = None
    if fcol:
        fval = st.sidebar.selectbox('篩選值', [''] + sorted(df[fcol].dropna().astype(str).unique().tolist())) or None
    sumc = st.sidebar.selectbox('加總欄', df.columns.tolist())
    if st.sidebar.button('生成 SUMIF 結果'):
        title = 'sumif_result'
        ws = wb[title] if title in wb.sheetnames else wb.create_sheet(title)
        first_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
        used = [c.column for c in first_row if c.value]
        start = 1 if not used else max(used) + 2
        _, cnt = generate_count_and_sum_sheet(ws, df, crit1, crit2, sumc, data_sheet, start, fcol, fval)
        st.success(f'已插入 SUMIF 公式，共 {cnt} 筆')
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        st.download_button('下載結果', buf, 'result_通用公式.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
elif mode == 'VLOOKUP 生成':
    key_sheet = st.sidebar.selectbox('查找工作表', sheets)
    df_key = pd.read_excel(BytesIO(bytes_data), sheet_name=key_sheet)
    key_col = st.sidebar.selectbox('查找欄', df_key.columns.tolist())
    subset_col = st.sidebar.selectbox('子集欄(可空)', [''] + df_key.columns.tolist())
    subset_val = '' if not subset_col else st.sidebar.selectbox('子集值', [''] + sorted(df_key[subset_col].dropna().astype(str).unique().tolist()))
    table_sheet = st.sidebar.selectbox('比對工作表', sheets)
    df_table = pd.read_excel(BytesIO(bytes_data), sheet_name=table_sheet)
    start_col = st.sidebar.selectbox('起始欄', df_table.columns.tolist())
    start_idx = df_table.columns.tolist().index(start_col)
    avail = df_table.columns.tolist()[start_idx:]
    out_cols = st.sidebar.multiselect('擷取欄(留空全選)', options=avail)
    out_cols = avail if not out_cols else out_cols
    if st.sidebar.button('生成 VLOOKUP 結果'):
        if subset_col and subset_val:
            title = f"vlookup_{key_sheet}_{subset_col}_{subset_val}"
        else:
            title = f"vlookup_{key_sheet}"
        ws = wb[title] if title in wb.sheetnames else wb.create_sheet(title)
        generate_vlookup(ws, df_key, key_sheet, key_col, subset_col, subset_val, df_table, table_sheet, start_idx, out_cols, blue_fill)
        st.success('已插入 VLOOKUP 公式')
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        st.download_button('下載結果', buf, 'result_通用公式.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
elif mode == '多欄合一':
    source = st.sidebar.selectbox('連結工作表', sheets)
    df_src = pd.read_excel(BytesIO(bytes_data), sheet_name=source)
    link_cols = st.sidebar.multiselect('擷取欄 (2~4 個)', options=df_src.columns.tolist())
    if st.sidebar.button('生成連結結果'):
        if len(link_cols) < 2 or len(link_cols) > 4:
            st.error('請選擇 2~4 個欄位')
        else:
            title = generate_link_sheet(wb, source, link_cols, df_src)
            st.success(f'已生成工作表 {title}')
            buf = BytesIO(); wb.save(buf); buf.seek(0)
            st.download_button('下載結果', buf, 'result_通用公式.xlsx', 'application/vnd.openxmlformats-officedocument-spreadsheetml.sheet')
else:  # SUM 檢測
    if st.sidebar.button('開始 SUM 檢測'):
        total = 0
        for sh in sheets:
            ws = wb[sh]
            df = pd.read_excel(BytesIO(bytes_data), sheet_name=sh)
            r_n, c_n = df.shape
            for c in range(c_n):
                for sj in range(c): total += apply_span_formula_horizontal(df, ws, sj, c-1, c, yellow_fill)
                for ej in range(c+1, c_n): total += apply_span_formula_horizontal(df, ws, c+1, ej, c, yellow_fill)
            for r in range(r_n):
                for si in range(r): total += apply_span_formula_vertical(df, ws, si, r-1, r, yellow_fill)
                for ei in range(r+1, r_n): total += apply_span_formula_vertical(df, ws, r+1, ei, r, yellow_fill)
        st.success(f'已插入 SUM 公式，共 {total} 筆')
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        st.download_button('下載結果', buf, 'result_通用公式.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
